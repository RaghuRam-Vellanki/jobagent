"""Write the Daily Brief Excel file.

Produces a 2-sheet workbook:
  Sheet 1 "Jobs": 12 columns per row, frozen header, hyperlinked apply link,
                  conditional color on Match Score (green/yellow/red).
  Sheet 2 "Summary": jobs per source + generation timestamp.

Caller hands us a list of `row` dicts and an output path.
"""
from __future__ import annotations
import logging
import os
from datetime import datetime
from collections import Counter

logger = logging.getLogger("brief.xlsx")

COLUMNS = [
    ("Source", "platform", 12),
    ("Company", "company", 28),
    ("Job Title", "title", 36),
    ("Location", "location", 22),
    ("Experience", "experience", 14),
    ("Salary", "salary", 18),
    ("Funding Status", "funding_status", 16),
    ("Company Size", "size_band", 14),
    ("Valuation", "valuation", 16),
    ("Match Score", "match_score", 12),
    ("Posted", "posted_at_source", 14),
    ("Apply Link", "url", 50),
]


def write_brief_xlsx(rows: list[dict], out_path: str) -> str:
    """Returns the absolute output path. Raises on openpyxl import error."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0071e3")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="DDDDDD")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for i, (label, _, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border
        ws.column_dimensions[c.column_letter].width = width

    # Data rows
    for r, row in enumerate(rows, start=2):
        for i, (_, key, _) in enumerate(COLUMNS, start=1):
            val = row.get(key, "")
            if key == "match_score":
                try:
                    val = round(float(val or 0))
                except Exception:
                    val = 0
            elif key == "posted_at_source" and isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            elif key == "url" and val:
                cell = ws.cell(row=r, column=i, value="Apply")
                cell.hyperlink = val
                cell.font = Font(color="0071e3", underline="single")
                cell.border = cell_border
                continue
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=(key in ("title", "company")))

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{ws.cell(row=len(rows) + 1, column=len(COLUMNS)).coordinate}"

    # Conditional formatting on Match Score column (col 10)
    if rows:
        score_range = f"J2:J{len(rows) + 1}"
        green = PatternFill("solid", fgColor="C6F6D5")
        yellow = PatternFill("solid", fgColor="FEFCBF")
        red = PatternFill("solid", fgColor="FED7D7")
        ws.conditional_formatting.add(score_range, CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=green))
        ws.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["40", "69"], fill=yellow))
        ws.conditional_formatting.add(score_range, CellIsRule(operator="lessThan", formula=["40"], fill=red))

    # Summary sheet
    s = wb.create_sheet("Summary")
    s["A1"] = "Daily Brief Summary"
    s["A1"].font = Font(bold=True, size=14)
    s["A3"] = "Generated"
    s["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M IST")
    s["A4"] = "Total jobs"
    s["B4"] = len(rows)
    s["A6"] = "By source"
    s["A6"].font = Font(bold=True)
    by_source = Counter(r.get("platform", "?") for r in rows)
    for i, (plat, count) in enumerate(sorted(by_source.items(), key=lambda x: -x[1]), start=7):
        s.cell(row=i, column=1, value=plat)
        s.cell(row=i, column=2, value=count)
    s.column_dimensions["A"].width = 22
    s.column_dimensions["B"].width = 14

    wb.save(out_path)
    logger.info(f"[xlsx] wrote {len(rows)} rows → {out_path}")
    return os.path.abspath(out_path)
